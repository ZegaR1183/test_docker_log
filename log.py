import os
import logging
from dotenv import load_dotenv
import traceback

# Загружаем переменные из .env файла
load_dotenv()
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
import psycopg2
from sqlalchemy import create_engine
from openpyxl import styles


# Настройка логирования и вывода в файл
LOG_DIR = "./logs"
LOG_FILE = f"{LOG_DIR}/app.log"

OUT_DIR = "./output_files"
IN_DIR = "./input_file"


# Создаём директорию для логов и файла excel, если нет
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(IN_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Константы
LOG_FILE_IN = f"{IN_DIR}/result_output"
LOG_FILE_OUT = "./temp/clear_log.txt"
DATA_FILE_OUT = "./temp/dict_all.txt"

# Количество полей для разных типов устройств
FIELDS_MX = 12
FIELDS_ACX_4000 = 8
FIELDS_ACX_2100 = 4

# Настройки вывода для DataFrame
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 300)

# Ключи для разных типов устройств
KEYS_MX = ['name', 'type', 'temp PEM_0', 'temp PEM_1', 'temp RE_0', 'temp RE_1',
           's_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5', 'date']
KEYS_ACX_4000 = ['name', 'type', 'temp PEM_0', 'temp PEM_1', 'temp RE_0', 's_fan_1', 's_fan_2', 'date']
KEYS_ACX_2100 = ['name', 'type', 'temp RE_0', 'date']

# Загрузка конфигурации из переменных окружения
def load_db_config() -> Dict[str, str]:
    """Загружает конфигурацию БД из переменных окружения."""
    db_config = {
        'dbname': os.getenv('POSTGRES_DB', 'postgres'),
        'user': os.getenv('POSTGRES_USER', 'postgres'),
        'password': os.getenv('POSTGRES_PASSWORD', ''),
        'host': os.getenv('DB_HOST', 'localhost'),
        'port': os.getenv('DB_PORT', '5432')
    }

    # Если пароль не передан - ошибка
    if not db_config['password']:
        raise ValueError("Пароль PostgreSQL не задан. Используйте переменную окружения POSTGRES_PASSWORD")

    return db_config

db_config = load_db_config()

# Подключение к PostgreSQL
DB_CONFIG = {
    'dbname': db_config['dbname'],
    'user': db_config['user'],
    'password': db_config['password'],
    'host': db_config['host'],
    'port': db_config['port']
}

# Создание движка SQLAlchemy для to_sql
engine = create_engine(f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@"
                       f"{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}")


def clear_log() -> None:
    """Очищает входной лог-файл и создает очищенный файл."""
    logger.info(f"Начало очистки лога из {LOG_FILE_IN}")
    with open(LOG_FILE_IN, "r", encoding="utf-8") as f_in, open(LOG_FILE_OUT, "w", encoding="utf-8") as f_out:
        for line in f_in:
            if "-----Outputs" in line:
                f_out.write(line)
            elif "Chassis" in line and "|match" not in line:
                parts = line.split()
                if len(parts) >= 3:
                    f_out.write(parts[2] + "\n")
            elif "Temp  PEM 0" in line:
                parts = line.split()
                if len(parts) >= 5:
                    f_out.write(parts[4] + "\n")
            elif "PEM " in line and "|match" not in line:
                parts = line.split()
                if len(parts) >= 4:
                    f_out.write(parts[3] + "\n")
            elif "Routing Engine" in line and "CPU" not in line:
                parts = line.split()
                if len(parts) >= 4:
                    idx = 3 if len(parts) == 10 else 4
                    f_out.write(parts[idx] + "\n")
            elif "Fan" in line:
                parts = line.split()
                if len(parts) >= 4:
                    f_out.write(parts[3] + "\n")
            elif "Current time" in line:
                parts = line.split()
                if len(parts) >= 3:
                    timestamp_part = parts[2].split('+')[0]
                    f_out.write(timestamp_part + "\n")
    logger.info(f"Лог очищен и сохранён в {LOG_FILE_OUT}")


def read_clear_data() -> List[Dict[str, Any]]:
    """Читает очищенный файл и преобразует данные в список словарей."""
    try:
        with open(LOG_FILE_OUT, "r", encoding="utf-8") as file:
            lines = [line.strip() for line in file.readlines()]

        list_all: List[List[str]] = []
        current_device: Optional[str] = None

        for line in lines:
            if "-----Outputs from " in line:
                current_device = line.split()[-2]
                list_all.append([current_device])
            elif current_device and line:
                list_all[-1].append(line)

        list_dict: List[Dict[str, Any]] = []
        for item in list_all:
            item_len = len(item)
            if item_len == FIELDS_MX:
                list_dict.append(dict(zip(KEYS_MX, item)))
            elif item_len == FIELDS_ACX_4000:
                list_dict.append(dict(zip(KEYS_ACX_4000, item)))
            elif item_len == FIELDS_ACX_2100:
                list_dict.append(dict(zip(KEYS_ACX_2100, item)))
            else:
                logger.warning(f"Пропуск записи с неожиданным количеством полей: {item_len}")

        logger.info(f"Прочитано записей: {len(list_dict)}")
        return list_dict

    except FileNotFoundError:
        logger.error(f"Файл {LOG_FILE_OUT} не найден")
        return []
    except Exception as e:
        logger.error(f"Ошибка при чтении файла: {e}")
        return []


def create_df(list_dict: List[Dict[str, Any]]) -> pd.DataFrame:
    """Создаёт DataFrame и нормализует названия колонок."""
    df = pd.DataFrame(list_dict)

    # Переименование колонок с пробелами
    df = df.rename(columns={
        'temp PEM_0': 'temp_pem_0',
        'temp PEM_1': 'temp_pem_1',
        'temp RE_0': 'temp_re_0',
        'temp RE_1': 'temp_re_1'
    })

    # Приведение типов
    numeric_cols = ['temp_pem_0', 'temp_pem_1', 'temp_re_0', 'temp_re_1',
                    's_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')

    if 'date' in df.columns:
        df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.date

    logger.info(f"DataFrame создан: {len(df)} строк, {len(df.columns)} колонок")
    return df


def create_table_if_not_exists() -> None:
    """Создаёт таблицу devices, если она не существует."""
    create_table_query = """
    CREATE TABLE IF NOT EXISTS devices (
        id SERIAL PRIMARY KEY,
        name VARCHAR(100),
        type VARCHAR(50),
        temp_pem_0 INTEGER,
        temp_pem_1 INTEGER,
        temp_re_0 INTEGER,
        temp_re_1 INTEGER,
        s_fan_1 INTEGER,
        s_fan_2 INTEGER,
        s_fan_3 INTEGER,
        s_fan_4 INTEGER,
        s_fan_5 INTEGER,
        date DATE
    );
    """
    try:
        with psycopg2.connect(**DB_CONFIG) as conn:
            with conn.cursor() as cur:
                cur.execute(create_table_query)
                conn.commit()
        logger.info("Таблица 'devices' проверена (создана или уже существует)")
    except Exception as e:
        logger.error(f"Ошибка при создании таблицы: {e}")
        raise


def save_to_postgres(df: pd.DataFrame, replace: bool = False) -> None:
    """Сохраняет DataFrame в таблицу devices."""
    try:
        if_exists = 'replace' if replace else 'append'
        df.to_sql('devices', engine, if_exists=if_exists, index=False, method='multi')
        logger.info(f"Данные сохранены в PostgreSQL (записей: {len(df)}, режим: {if_exists})")
    except Exception as e:
        logger.error(f"Ошибка при сохранении в PostgreSQL: {e}")


def analyze_data(df: pd.DataFrame) -> dict:
    """Анализирует данные и возвращает статистику."""
    cnt_chassi = df.groupby('type', as_index=False).size()

    # Проверка существования колонок вентиляторов
    fan_columns = [c for c in ['s_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5'] if c in df.columns]
    if fan_columns:
        cnt_1_fan_check = (df[fan_columns] == 0).any(axis=1).sum()
        df_fan_alarm = df[(df[fan_columns] == 0).any(axis=1)]
    else:
        cnt_1_fan_check = 0
        df_fan_alarm = pd.DataFrame()

    # Проверка двух отключённых вентиляторов (только если есть обе колонки)
    cnt_2_fans_check = 0
    if 's_fan_1' in df.columns and 's_fan_2' in df.columns:
        cnt_2_fans_check = ((df['s_fan_1'] == 0) & (df['s_fan_2'] == 0) & (df["type"] != "MX104")).sum()

    # Проверка температурных колонок
    temp_columns = [c for c in ['temp_pem_0', 'temp_pem_1', 'temp_re_0', 'temp_re_1'] if c in df.columns]
    if temp_columns:
        temp_warm = ((df[temp_columns] > 50).any(axis=1)).sum()
        df_temp_alarm = df[(df[temp_columns] > 50).any(axis=1)]
    else:
        temp_warm = 0
        df_temp_alarm = pd.DataFrame()

    logger.info("Анализ данных завершён")
    return {
        'device_count': cnt_chassi,
        'fans_disabled': cnt_1_fan_check,
        'two_fans_disabled': cnt_2_fans_check,
        'high_temp_devices': temp_warm,
        'fan_alarm_devices': df_fan_alarm,
        'temp_alarm_devices': df_temp_alarm
    }

def save_to_excel_sheets(df: pd.DataFrame, analysis_results: dict) -> None:
    """Сохраняет данные в Excel файл с отдельными листами."""
    current_date = datetime.now().strftime("%Y-%m-%d")
    output_file = f"{OUT_DIR}/alarm_report_metro {current_date}.xlsx"

    try:
        df_excel = df.drop(columns=['date'], errors='ignore')

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_excel.to_excel(writer, sheet_name="Общая информация", index=False)

            pd.DataFrame([{
                'Устройства с одним отключенным вентилятором': analysis_results['fans_disabled'],
                'Устройства с двумя отключенными вентиляторами': analysis_results['two_fans_disabled'],
                'Устройства с температурой > 50': analysis_results['high_temp_devices']
            }]).to_excel(writer, sheet_name='Статистика', index=False)

            df_fans = analysis_results['fan_alarm_devices'].drop(columns=['date'], errors='ignore')
            if not df_fans.empty:
                df_fans.to_excel(writer, sheet_name='Вентиляторы', index=False)

            df_temps = analysis_results['temp_alarm_devices'].drop(columns=['date'], errors='ignore')
            if not df_temps.empty:
                df_temps.to_excel(writer, sheet_name='Температура', index=False)

            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                for column in worksheet.columns:
                    for cell in column:
                        cell.alignment = styles.Alignment(horizontal='center')

            worksheet = writer.sheets["Общая информация"]
            fan_columns = ['s_fan_1', 's_fan_2', 's_fan_3', 's_fan_4', 's_fan_5']
            for col in fan_columns:
                if col in df_excel.columns:
                    col_idx = df_excel.columns.get_loc(col) + 1
                    for row in range(2, len(df_excel) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value == 0:
                            cell.fill = styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            temp_columns = ['temp_pem_0', 'temp_pem_1', 'temp_re_0', 'temp_re_1']
            for col in temp_columns:
                if col in df_excel.columns:
                    col_idx = df_excel.columns.get_loc(col) + 1
                    for row in range(2, len(df_excel) + 2):
                        cell = worksheet.cell(row=row, column=col_idx)
                        try:
                            if cell.value is not None and float(cell.value) > 50:
                                cell.fill = styles.PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        except (ValueError, TypeError):
                            pass

        logger.info(f"Данные сохранены в {output_file}")
    except Exception as e:
        logger.error(f"Ошибка при сохранении в Excel: {e}")
        logger.error(f"Stack trace:\n{traceback.format_exc()}")  # 🔥 Важно!
        raise  # 🔥 Прервать выполнение, чтобы увидеть ошибку

if __name__ == "__main__":
    logger.info("Запуск скрипта")

    # Проверка и создание таблицы
    create_table_if_not_exists()

    # Очистка и чтение данных
    clear_log()
    data = read_clear_data()

    if not data:
        logger.warning("Нет данных для обработки.")
        exit()

    df = create_df(data)

    # Анализ данных
    analysis_results = analyze_data(df)

    logger.info(f"Устройства с отключёнными вентиляторами: {analysis_results['fans_disabled']}")
    logger.info(f"Устройства с двумя отключёнными вентиляторами (не MX104): {analysis_results['two_fans_disabled']}")
    logger.info(f"Устройства с температурой > 50°C: {analysis_results['high_temp_devices']}")

    # Сохранение в PostgreSQL (replace=True для перезаписи, False для дополнения)
    save_to_postgres(df, replace=False)

    # Сохранение в Excel
    save_to_excel_sheets(df, analysis_results)

    logger.info("Скрипт завершён")
